from google.cloud import documentai_v1beta2 as documentai
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def parse_table(project_id="table-extract-304916", input_uri="gs://test-data-table-extract/test2.pdf",):
    """Parse a form"""
    client = documentai.DocumentUnderstandingServiceClient()
    gcs_source = documentai.types.GcsSource(uri=input_uri)

    # mime_type can be application/pdf, image/tiff,
    # and image/gif, or application/json
    input_config = documentai.types.InputConfig(gcs_source=gcs_source, mime_type="application/pdf")

    # Improve table parsing results by providing bounding boxes
    # specifying where the box appears in the document (optional)
    table_bound_hints = [
        documentai.types.TableBoundHint(
            page_number=1,
            bounding_box=documentai.types.BoundingPoly(
                # Define a polygon around tables to detect
                # Each vertice coordinate must be a number between 0 and 1
                normalized_vertices=[
                    # Top left
                    documentai.types.geometry.NormalizedVertex(x=0, y=0),
                    # Top right
                    documentai.types.geometry.NormalizedVertex(x=1, y=0),
                    # Bottom right
                    documentai.types.geometry.NormalizedVertex(x=1, y=1),
                    # Bottom left
                    documentai.types.geometry.NormalizedVertex(x=0, y=1),
                ]
            ),
        )
    ]

    # Setting enabled=True enables form extraction
    table_extraction_params = documentai.types.TableExtractionParams(enabled=True, table_bound_hints=table_bound_hints)

    # Location can be 'us' or 'eu'
    parent = "projects/{}/locations/us".format(project_id)
    request = documentai.types.ProcessDocumentRequest(parent=parent, input_config=input_config, table_extraction_params=table_extraction_params, )

    document = client.process_document(request=request)

    def _get_text(el):
        """Convert text offset indexes into text snippets."""
        response = ""
        # If a text segment spans several lines, it will
        # be stored in different text segments.
        for segment in el.text_anchor.text_segments:
            start_index = segment.start_index
            end_index = segment.end_index
            response += document.text[start_index:end_index]
        return response.replace("\n", "")

    wb = Workbook()

    dest_fileName = 'test.xlsx'

    ws1 = wb.active
    ws1.title = "Data"

    for page in document.pages:
        col = 1
        row1 = 1
        print("Page number: {}".format(page.page_number))
        for table_num, table in enumerate(page.tables):
            print("Table {}: ".format(table_num))
            for row_num, row in enumerate(table.header_rows):
                cells = "\t".join([_get_text(cell.layout) for cell in row.cells])
                print("Header Row {}: {}".format(row_num, cells))
                for cell in row.cells:
                    val = _get_text(cell.layout)
                    ws1.cell(row=row1, column=col, value=val)
                    print(val + "\t")
                    col = col + 1
                col = 1
            for row_num, row in enumerate(table.body_rows):
                cells = "\t".join([_get_text(cell.layout) for cell in row.cells])
                print("Row {}: {}".format(row_num, cells))
                row1 = row1 + 1
                col = 1
                for cell in row.cells:
                    val = _get_text(cell.layout)
                    ws1.cell(row=row1, column=col, value=val)
                    print(val + "\t")
                    col = col + 1

    wb.save(filename=dest_fileName)
parse_table()
